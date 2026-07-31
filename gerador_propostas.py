import streamlit as st
import pandas as pd
import numpy as np
import re
import os
import io
import base64
import unicodedata
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Tenta importar as bibliotecas de PDF
try:
    import weasyprint
    HAS_PDF_GENERATOR = True
except ImportError:
    HAS_PDF_GENERATOR = False

st.set_page_config(page_title="Gerador de Propostas - Leves", layout="wide")

st.title("📦 Gerador de Propostas e Movimentação de Leves")
st.markdown("Bem-vindo! Faça o upload das bases extraídas do Looker para começar.")

# --- BARRA LATERAL: UPLOADS E LINKS LOOKER ---
with st.sidebar.expander("1. Upload de Bases de Dados", expanded=True):
    st.markdown("**Tabelas frete peso praticadas**")
    st.markdown("[Link Looker: 26300](https://loggi.looker.com/looks/26300)")
    file_frete = st.file_uploader("Upload Frete Peso", type=["xlsx", "csv"], label_visibility="collapsed")
    
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**Abrangências atuais**")
    st.markdown("[Link Looker: 26301](https://loggi.looker.com/looks/26301)")
    file_abrangencia = st.file_uploader("Upload Abrangência", type=["xlsx", "csv"], label_visibility="collapsed")
    
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**SLOs globais das cidades**")
    st.markdown("[Link Looker: 26303](https://loggi.looker.com/looks/26303)")
    file_slos = st.file_uploader("Upload SLOs", type=["xlsx", "csv"], label_visibility="collapsed")
    
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**Volume de pacotes (30 dias)**")
    st.markdown("[Link Looker: 26302](https://loggi.looker.com/looks/26302)")
    file_volume = st.file_uploader("Upload Volume", type=["xlsx", "csv"], label_visibility="collapsed")

st.sidebar.markdown("<br><hr><div style='text-align: center;'><small>Desenvolvido por <b>Matheus Zanetti</b></small></div>", unsafe_allow_html=True)

# --- INICIALIZAÇÃO DE ESTADOS ---
if "num_cenarios" not in st.session_state:
    st.session_state["num_cenarios"] = 1

# Cores pastéis para organizar os cenários visualmente (C1, C2, C3, C4...)
CORES_CENARIOS = ['#e6f2ff', '#e6ffe6', '#fff2e6', '#f2e6ff', '#ffe6e6']

# --- FUNÇÕES DE TRATAMENTO E PADRONIZAÇÃO DE DADOS ---
@st.cache_data
def load_data(file):
    if file.name.endswith('.csv'):
        return pd.read_csv(file)
    return pd.read_excel(file)

@st.cache_data
def load_local_excel(filename):
    if not os.path.exists(filename):
        return None
    return pd.read_excel(filename)

def normalize_string(s):
    if pd.isna(s): return ""
    s = str(s).lower().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def padronizar_colunas_frete(df):
    mapa = {
        "Leve Contract Region LMC name": "LMC name",
        "Leve Contract Region table name": "table name",
        "Leve Contract Region label": "label",
        "Leve Contract Region on time amount": "on time amount",
        "Leve Contract Region out of time amount": "out of time amount",
        "Leve Contract Region service type": "service type",
        "Leve Contract Region Faixa de peso (g/m³)": "Faixa de peso cubado (g)",
        "Faixa de peso (g/m³)": "Faixa de peso cubado (g)"
    }
    return df.rename(columns=mapa)

def padronizar_colunas_abrangencia(df):
    mapa = {
        "Territorial Scope Pricing Regions LMC Name": "LMC Name",
        "Territorial Scope Pricing Regions Pricing Region": "Região de preço 2023",
        "Territorial Scope Pricing Regions City": "Cidade",
        "Territorial Scope Pricing Regions State": "State",
        "Territorial Scope Pricing Regions Service Type": "Tipo de serviço",
        "Territorial Scope Pricing Regions SLO Lastmile": "Prazo adicional"
    }
    return df.rename(columns=mapa)

def padronizar_colunas_slos(df):
    mapa = {
        "Territorial Scope Pricing Regions City": "Cidade",
        "Territorial Scope Pricing Regions Pricing Region": "Região de preço 2023",
        "Territorial Scope Pricing Regions State": "State",
        "Territorial Scope Pricing Regions Service Type": "Tipo de serviço",
        "Territorial Scope Pricing Regions SLO": "SLO"
    }
    return df.rename(columns=mapa)

def padronizar_colunas_volume(df):
    mapa = {
        "Package Charge Leve Last Mile Company Name": "Leve",
        "Distribution and Expedition Center Locations Routing Code": "Routing Code",
        "Package Charge Leve Region label": "Region label",
        "Package Charge Leve Region Label": "Region label",
        "Package Destination City": "Cidade",
        "Package Charge Leve Service Charge Type": "Service Charge Type",
        "Package Charge Leve # Packages": "# Total Packages",
        "Faixa pesos": "Faixa de peso cubado (g)",
        "Faixa Pesos": "Faixa de peso cubado (g)",
        "Package Charge Leve Faixa pesos": "Faixa de peso cubado (g)"
    }
    return df.rename(columns=mapa)

@st.cache_data
def processar_frete(df_frete):
    df_recentes = df_frete.drop_duplicates(subset=['LMC name'], keep='first')
    tabelas_validas = df_recentes[['LMC name', 'table name']]
    df_filtrado = df_frete.merge(tabelas_validas, on=['LMC name', 'table name'], how='inner')
    return df_filtrado

@st.cache_data
def processar_slos(df_slos):
    df_slos['prioridade'] = np.where(df_slos['Tipo de serviço'].str.contains('express', case=False, na=False), 1, 2)
    df_slos = df_slos.sort_values(by=['Cidade', 'prioridade', 'SLO'], ascending=[True, True, True])
    df_slos_clean = df_slos.drop_duplicates(subset=['Cidade'], keep='first')
    return df_slos_clean

@st.cache_data
def processar_nomes_leves(df_volume):
    mapping = df_volume[['Leve', 'Routing Code']].drop_duplicates().dropna()
    mapping['nome_completo'] = mapping['Leve'] + " (" + mapping['Routing Code'] + ")"
    return mapping

@st.cache_data
def processar_price_var(df_price):
    df_price.columns = ['Faixa de peso cubado (g)', 'Multiplicador']
    return df_price

def formatar_moeda(valor):
    try:
        return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return str(valor)

def extrair_estado(nome_leve):
    match = re.search(r'-\s*([A-Z]{2})\b', nome_leve)
    if match:
        return match.group(1)
    return None

# --- FUNÇÃO DE FORMATAÇÃO DE EXCEL ---
def formatar_excel(writer):
    workbook = writer.book
    font_padrao = Font(name='Inter', size=10)
    font_cabecalho = Font(name='Inter', size=10, bold=True)
    alinhamento = Alignment(horizontal='center', vertical='center', wrap_text=False)
    
    borda_cinza = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.sheet_view.showGridLines = False
        
        col_formats = {}
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val_str = cell.value.lower()
                    if "(r$)" in val_str or "tarifa" in val_str or "ticket" in val_str or "faturamento" in val_str or "impacto" in val_str or "fat." in val_str or "tk" in val_str:
                        col_formats[cell.column_letter] = '"R$" #,##0.00'
                    elif "(%)" in val_str or "% var" in val_str or "ajuste" in val_str or "% aum" in val_str:
                        col_formats[cell.column_letter] = '0.00%'
                    elif "pacotes" in val_str or "volume" in val_str:
                        col_formats[cell.column_letter] = '#,##0'

        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 18
            for cell in row:
                if cell.value is not None:
                    if cell.row == 1 or str(cell.value) in col_formats.keys() or "Total Geral" in str(row[0].value):
                        cell.font = font_cabecalho
                    else:
                        cell.font = font_padrao
                        
                    if cell.column_letter in col_formats and type(cell.value) in [int, float]:
                        cell.number_format = col_formats[cell.column_letter]
                            
                    cell.alignment = alinhamento
                    cell.border = borda_cinza
                    
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2) * 1.15
            if adjusted_width < 12: adjusted_width = 12
            ws.column_dimensions[column].width = adjusted_width


# --- GERADOR DE PDF ---
def generate_html_pdf(nome_destino, df_resumo_comparativo,
                      detalhes_cenarios, df_abrangencia_out, dict_tabelas_out, cenarios_nomes):
    fuso_brasilia = datetime.now(timezone(timedelta(hours=-3)))
    data_extracao = fuso_brasilia.strftime("%d/%m/%Y às %H:%M")

    def format_money(val):
        return f"R$ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    def format_perc(val):
        return f"{val:+.2f}%"

    logo_html = ""
    logo_path = "logo.png"
    if not os.path.exists(logo_path) and os.path.exists("logo.png.png"):
        logo_path = "logo.png.png"
        
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode()
            logo_html = f'<img src="data:image/png;base64,{encoded_string}" style="position: fixed; top: -20mm; right: 0; width: 45px; height: auto; z-index: 1000;">'

    html_content = f"""
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8">
        <title>Relatório de Simulação - Loggi</title>
        <style>
            @page {{
                size: A4 landscape;
                margin: 30mm 15mm 20mm 15mm;
                background-color: #ffffff;
                @bottom-center {{
                    content: "Simulador de Movimentação de Leves | Página " counter(page);
                    font-family: 'Montserrat', 'Helvetica Neue', Helvetica, Arial, sans-serif;
                    font-size: 8pt;
                    color: #888888;
                }}
            }}
            body {{
                font-family: 'Montserrat', 'Helvetica Neue', Helvetica, Arial, sans-serif;
                margin: 0;
                padding: 0;
                color: #000000;
                background-color: #ffffff;
                font-size: 9pt;
                line-height: 1.5;
            }}
            h1 {{ color: #002766; font-size: 16pt; text-align: center; margin-top: 10px; margin-bottom: 5px; }}
            h2 {{ color: #006aff; font-size: 12pt; border-bottom: 2px solid #00baff; padding-bottom: 4px; margin-top: 25px; margin-bottom: 15px; }}
            .header-meta {{ text-align: center; color: #666; font-size: 8pt; margin-bottom: 25px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 8pt; }}
            th {{ background-color: #002766; color: #ffffff; font-weight: bold; text-align: center; padding: 6px 4px; border: 1px solid #002766; }}
            td {{ padding: 5px 4px; border: 1px solid #e0e0e0; text-align: center; color: #333333; }}
            tr:nth-child(even) {{ background-color: #f4f8fb; }}
            .page-break {{ page-break-before: always; }}
            .highlight-row td {{ background-color: #002766 !important; color: #ffffff !important; font-weight: bold !important; }}
            .arrow-up-red {{ color: #ff4b4b; font-weight: bold; }}
            .arrow-down-green {{ color: #09ab3b; font-weight: bold; }}
            .arrow-neutral {{ color: #888888; font-weight: bold; }}
        </style>
    </head>
    <body>
    {logo_html}
        <h1>Relatório Comparativo de Cenários</h1>
        <div class="header-meta">
            Gerado em: {data_extracao} | Destino: {nome_destino}
        </div>
    """

    def generate_html_table_styled(df):
        if df is None or df.empty: return "<p>Sem dados.</p>"
        html = "<table><thead><tr>"
        for col in df.columns: html += f"<th>{col}</th>"
        html += "</tr></thead><tbody>"
        for _, row in df.iterrows():
            if str(row.iloc[0]).lower() == 'total geral':
                html += f"<tr class='highlight-row'>"
                for val in row:
                    html += f"<td>{val}</td>"
                html += "</tr>"
            else:
                html += "<tr>"
                for col_idx, val in enumerate(row):
                    col_name = df.columns[col_idx]
                    bg_color = ""
                    for i, cen in enumerate(cenarios_nomes):
                        if cen in col_name or f"Cenário {i+1}" in col_name:
                            bg_color = f" background-color: {CORES_CENARIOS[i % len(CORES_CENARIOS)]};"
                    html += f"<td style='{bg_color}'>{val}</td>"
                html += "</tr>"
        html += "</tbody></table>"
        return html

    html_content += f"""<h2>1. RESUMO COMPARATIVO DE CENÁRIOS</h2>"""
    
    df_resumo_html = df_comparativo.copy()
    for c in df_resumo_html.columns:
        if "Fat" in c or "Ticket" in c or "TK" in c or "Impacto" in c:
            df_resumo_html[c] = df_resumo_html[c].apply(lambda x: format_money(x) if pd.notna(x) and isinstance(x, (int, float)) else x)
        elif "%" in c or "Aum" in c:
            df_resumo_html[c] = df_resumo_html[c].apply(lambda x: format_perc(x * 100) if pd.notna(x) and isinstance(x, (int, float)) else x)
        elif "Vol" in c:
            df_resumo_html[c] = df_resumo_html[c].apply(lambda x: f"{int(x):,}".replace(",", ".") if pd.notna(x) else "-")

    html_content += generate_html_table_styled(df_resumo_html)

    def generate_simple_html_table(df):
        if df is None or df.empty: return "<p>Sem dados.</p>"
        html = "<table><thead><tr>"
        for col in df.columns: html += f"<th>{col}</th>"
        html += "</tr></thead><tbody>"
        for _, row in df.iterrows():
            html += "<tr>"
            for val in row: html += f"<td>{val}</td>"
            html += "</tr>"
        html += "</tbody></table>"
        return html

    html_content += f"""
        <div class="page-break"></div>
        <h2>2. ABRANGÊNCIA COMPLETA PROJETADA</h2>
    """
    colunas_abr_pdf = [c for c in df_abrangencia_out.columns if c != 'State']
    html_content += generate_simple_html_table(df_abrangencia_out[colunas_abr_pdf])

    html_content += f"""
        <div class="page-break"></div>
        <h2>3. TABELAS FRETE PESO PROJETADAS POR CENÁRIO</h2>
    """
    for cenario_nome, df_tabela in dict_tabelas_out.items():
        html_content += f"<h3>{cenario_nome}</h3>"
        df_tab_html = df_tabela.copy()
        df_tab_html['Valor dentro do prazo'] = df_tab_html['Valor dentro do prazo'].apply(format_money)
        df_tab_html['Valor fora do prazo'] = df_tab_html['Valor fora do prazo'].apply(format_money)
        html_content += generate_simple_html_table(df_tab_html)

    html_content += """
    </body>
    </html>
    """
    
    pdf_bytes = weasyprint.HTML(string=html_content).write_pdf()
    return pdf_bytes

# --- FLUXO PRINCIPAL ---
df_price_var_raw = load_local_excel("Price variation.xlsx")

if file_frete and file_abrangencia and file_slos and file_volume:
    if df_price_var_raw is None:
        st.error("Erro: O arquivo 'Price variation.xlsx' não foi encontrado. Por favor, certifique-se de que ele foi subido para o repositório do GitHub.")
    else:
        with st.spinner("Carregando e processando bases..."):
            df_frete = load_data(file_frete)
            df_abrangencia = load_data(file_abrangencia)
            df_slos = load_data(file_slos)
            df_volume = load_data(file_volume)
            
            df_frete = padronizar_colunas_frete(df_frete)
            df_abrangencia = padronizar_colunas_abrangencia(df_abrangencia)
            df_slos = padronizar_colunas_slos(df_slos)
            df_volume = padronizar_colunas_volume(df_volume)
            
        if 'Leve' not in df_volume.columns or 'Routing Code' not in df_volume.columns:
            st.error("🚨 **Colunas básicas ausentes no arquivo de Volume!**")
            st.stop()
            
        if 'Cidade' not in df_volume.columns or 'Faixa de peso cubado (g)' not in df_volume.columns:
            st.error("🚨 **Atenção: Base de Volume Desatualizada ou Incorreta!**")
            st.markdown("Por favor, verifique se a base extraída possui as colunas de **Cidade** e **Faixa Pesos**.")
            st.stop()
            
        with st.spinner("Finalizando processamento..."):
            df_volume['Cidade_Normalizada'] = df_volume['Cidade'].apply(normalize_string)
            df_abrangencia['Cidade_Normalizada'] = df_abrangencia['Cidade'].apply(normalize_string)
            
            df_price_var_clean = processar_price_var(df_price_var_raw)
            df_frete_clean = processar_frete(df_frete)
            df_slos_clean = processar_slos(df_slos)
            
            df_nomes_leves = processar_nomes_leves(df_volume)
            
            df_volume['Faixa de peso cubado (g)'] = df_volume['Faixa de peso cubado (g)'].astype(str).str.strip()
            df_volume_grouped = df_volume.groupby(
                ['Leve', 'Cidade_Normalizada', 'Faixa de peso cubado (g)'],
                as_index=False
            ).agg({
                '# Total Packages': 'sum',
                'Cidade': 'first'
            })
            df_volume = df_volume_grouped
            
        st.sidebar.success("Todas as bases carregadas e padronizadas!")
    
        with st.expander("2. Seleção de Leves", expanded=True):
            leves_disponiveis = df_frete_clean['LMC name'].dropna().unique().tolist()
            
            mapa_nomes = {}
            mapa_routing = {} 
            for lmc in leves_disponiveis:
                match = df_nomes_leves[df_nomes_leves['Leve'] == lmc]
                if not match.empty:
                    nome_formatado = match['nome_completo'].values[0]
                    mapa_nomes[nome_formatado] = lmc
                    mapa_routing[lmc] = match['Routing Code'].values[0]
                else:
                    mapa_nomes[lmc] = lmc
                    mapa_routing[lmc] = "-"
                    
            lista_nomes_exibicao = list(mapa_nomes.keys())
            leves_selecionados_formatados = st.multiselect("Selecione os Leves envolvidos na negociação:", lista_nomes_exibicao)
            leves_selecionados = [mapa_nomes[nome] for nome in leves_selecionados_formatados]
    
        if leves_selecionados:
            with st.expander("3. Definição das Cidades Base", expanded=True):
                cidades_base_dict = {}
                cols = st.columns(len(leves_selecionados))
                for idx, leve in enumerate(leves_selecionados):
                    nome_exibicao = leves_selecionados_formatados[idx]
                    estado_do_leve = extrair_estado(leve)
                    if estado_do_leve:
                        df_cidades_estado = df_slos_clean[df_slos_clean['State'] == estado_do_leve]
                    else:
                        df_cidades_estado = df_slos_clean 
                        
                    opcoes_cidades = {}
                    for _, row in df_cidades_estado.iterrows():
                        cid = str(row['Cidade'])
                        est = str(row['State'])
                        opcoes_cidades[f"{cid} - {est}"] = cid
                        
                    with cols[idx]:
                        cidade_escolhida_display = st.selectbox(f"Cidade Base para:\n{nome_exibicao}", sorted(list(opcoes_cidades.keys())), key=f"cidade_{leve}")
                        cidades_base_dict[leve] = opcoes_cidades[cidade_escolhida_display]
    
            with st.expander("4. Dados Atuais dos Leves Selecionados", expanded=False):
                df_abrangencia.rename(columns={'Região de preço 2023': 'Região de preço'}, inplace=True)
                tab1, tab2 = st.tabs(["Tabela Frete Peso Atual", "Abrangência e Prazos"])
                with tab1:
                    for idx, leve in enumerate(leves_selecionados):
                        st.subheader(f"Tabela Frete Peso: {leves_selecionados_formatados[idx]}")
                        df_frete_leve = df_frete_clean[df_frete_clean['LMC name'] == leve].copy()
                        df_frete_leve['Routing Code'] = mapa_routing.get(leve, "-")
                        df_frete_leve.rename(columns={'label': 'Região de preço', 'on time amount': 'Valor do pacote dentro do prazo', 'out of time amount': 'Valor do pacote fora do prazo'}, inplace=True)
                        df_frete_leve['Valor do pacote dentro do prazo'] = df_frete_leve['Valor do pacote dentro do prazo'].apply(formatar_moeda)
                        df_frete_leve['Valor do pacote fora do prazo'] = df_frete_leve['Valor do pacote fora do prazo'].apply(formatar_moeda)
                        col_exib = ['LMC name', 'Routing Code', 'Região de preço', 'Faixa de peso cubado (g)', 'Valor do pacote dentro do prazo', 'Valor do pacote fora do prazo', 'table name']
                        st.dataframe(df_frete_leve[[c for c in col_exib if c in df_frete_leve.columns]], use_container_width=True, hide_index=True)
                with tab2:
                    for idx, leve in enumerate(leves_selecionados):
                        st.subheader(f"Abrangência e Prazos: {leves_selecionados_formatados[idx]}")
                        df_abrangencia_leve = df_abrangencia[df_abrangencia['LMC Name'] == leve].copy()
                        df_abrangencia_leve['Routing Code'] = mapa_routing.get(leve, "-")
                        df_abrangencia_leve['SLO Local (Arquivo)'] = df_abrangencia_leve['Prazo adicional']
                        st.dataframe(df_abrangencia_leve[['LMC Name', 'Routing Code', 'Região de preço', 'Cidade', 'State', 'SLO Local (Arquivo)']], use_container_width=True, hide_index=True)
                        
            with st.expander("5. Definição do Leve/Lead de Destino", expanded=True):
                tipo_destino = st.radio("O destino da movimentação será para:", ["Um Leve Existente (já selecionado)", "Um Novo Lead"])
                nome_destino_final = ""
                cidade_base_destino = ""
                if tipo_destino == "Um Leve Existente (já selecionado)":
                    nome_destino_display = st.selectbox("Selecione o Leve de Destino:", leves_selecionados_formatados)
                    nome_destino_final = mapa_nomes.get(nome_destino_display)
                    cidade_base_destino = cidades_base_dict.get(nome_destino_final)
                    st.info(f"**Cidade Base do Destino:** {cidade_base_destino}")
                else:
                    col_n1, col_n2, col_n3 = st.columns(3)
                    with col_n1: nome_destino_final = st.text_input("Nome do Novo Lead:", placeholder="Ex: Lead - SP Sorocaba...")
                    with col_n2: estado_lead = st.selectbox("Estado do Novo Lead:", sorted(df_slos_clean['State'].dropna().unique().tolist()))
                    with col_n3: cidade_base_destino = st.selectbox("Cidade Base do Novo Lead:", sorted(df_slos_clean[df_slos_clean['State'] == estado_lead]['Cidade'].tolist()))
    
            if nome_destino_final and cidade_base_destino:
                with st.expander("6. Manipulação de Abrangência", expanded=True):
                    config_key = f"{','.join(leves_selecionados)}_{nome_destino_final}"
                    if "mov_config_key" not in st.session_state or st.session_state.mov_config_key != config_key:
                        df_abrangencia_alvo = df_abrangencia[df_abrangencia['LMC Name'].isin(leves_selecionados)].copy()
                        df_mov = df_abrangencia_alvo[['LMC Name', 'Região de preço', 'Cidade', 'State']].copy()
                        df_mov['Destino'] = "Manter no Leve Atual"
                        st.session_state.df_movimentacao = df_mov
                        st.session_state.mov_config_key = config_key

                    opcoes_destino = ["Manter no Leve Atual", nome_destino_final]
                    st.markdown("⚡ **Ações em Massa**")
                    col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns([2.5, 2.5, 2.5, 1.5, 1.5])
                    with col_m1: bulk_lmc = st.selectbox("1. Filtrar Origem:", ["Selecione..."] + leves_selecionados, key="bulk_lmc")
                    with col_m2:
                        opcoes_reg = ["Todas as Regiões"]
                        if bulk_lmc != "Selecione...": opcoes_reg += sorted(list(st.session_state.df_movimentacao[st.session_state.df_movimentacao['LMC Name'] == bulk_lmc]['Região de preço'].unique()))
                        bulk_reg = st.selectbox("2. Filtrar Região:", opcoes_reg, key="bulk_reg")
                    with col_m3: bulk_dest = st.selectbox("3. Escolher Destino:", opcoes_destino, key="bulk_dest")
                    with col_m4:
                        st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
                        def aplicar_em_massa():
                            l, r, d = st.session_state.bulk_lmc, st.session_state.bulk_reg, st.session_state.bulk_dest
                            if l != "Selecione...":
                                mask = st.session_state.df_movimentacao['LMC Name'] == l
                                if r != "Todas as Regiões": mask &= st.session_state.df_movimentacao['Região de preço'] == r
                                st.session_state.df_movimentacao.loc[mask, 'Destino'] = d
                        st.button("Aplicar Ação", on_click=aplicar_em_massa, use_container_width=True)
                    with col_m5:
                        st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
                        st.button("🔄 Resetar", on_click=lambda: st.session_state.df_movimentacao.assign(Destino="Manter no Leve Atual"), use_container_width=True)

                    df_editado = st.data_editor(
                        st.session_state.df_movimentacao,
                        column_config={"Destino": st.column_config.SelectboxColumn("Destino", options=opcoes_destino, required=True)},
                        disabled=["LMC Name", "Região de preço", "Cidade", "State"], hide_index=True, use_container_width=True, height=400
                    )
                    st.session_state.df_movimentacao = df_editado
                    df_movidos = df_editado[df_editado['Destino'] == nome_destino_final].copy()

                with st.expander("7. Estratégia de Precificação", expanded=True):
                    pode_prosseguir = False
                    estrategia_preco = None
                    tabela_base_selecionada = None
                    if df_movidos.empty and tipo_destino == "Um Novo Lead":
                        st.warning("Nenhum município foi movimentado para o Novo Lead ainda.")
                    elif df_movidos.empty and tipo_destino == "Um Leve Existente (já selecionado)":
                        st.info("💡 **Modo de Reajuste Comercial**")
                        estrategia_preco = "Reajuste da Tabela Atual (Sem movimentação)"
                        tabela_base_selecionada = nome_destino_final
                        pode_prosseguir = True
                    else:
                        estrategia_preco = st.radio("Estratégia:", ["Gerar Tabela Equivalente (Focada em manter Impacto Neutro)", "Utilizar uma Tabela Existente (Manter tabela do destino atual)"])
                        if estrategia_preco == "Utilizar uma Tabela Existente (Manter tabela do destino atual)":
                            if tipo_destino == "Um Leve Existente (já selecionado)":
                                tabela_base_selecionada = nome_destino_final
                            else:
                                nome_base_display = st.selectbox("Qual tabela atual usar como base?", leves_selecionados_formatados)
                                tabela_base_selecionada = mapa_nomes.get(nome_base_display)
                        pode_prosseguir = True
                
                # --- PROCESSAMENTO DOS RESULTADOS ---
                if pode_prosseguir:
                    
                    dict_tabelas_finais = {}
                    tabelas_atuais_pdf = {}
                    
                    if tipo_destino == "Um Leve Existente (já selecionado)":
                        df_abrangencia_existente = df_abrangencia[df_abrangencia['LMC Name'] == nome_destino_final].copy()
                        df_abrangencia_existente['Observação'] = "Abrangência Atual"
                    else:
                        df_abrangencia_existente = pd.DataFrame(columns=['LMC Name', 'Região de preço', 'Cidade', 'State', 'Observação'])

                    df_movidos_fmt = df_movidos.copy()
                    df_movidos_fmt['Observação'] = "Migrado de: " + df_movidos_fmt['LMC Name']
                    df_escopo_final = pd.concat([df_abrangencia_existente[['Cidade', 'State', 'Região de preço', 'Observação']], df_movidos_fmt[['Cidade', 'State', 'Região de preço', 'Observação']]], ignore_index=True).drop_duplicates(subset=['Cidade'])
                    
                    slo_base_dest = df_slos_clean[df_slos_clean['Cidade'] == cidade_base_destino]['SLO'].values
                    slo_base_dest_val = slo_base_dest[0] if len(slo_base_dest) > 0 else 0
                    
                    df_escopo_final = df_escopo_final.merge(df_slos_clean[['Cidade', 'SLO']], on='Cidade', how='left')
                    df_escopo_final['Novo SLO Local'] = df_escopo_final['SLO'] - slo_base_dest_val
                    df_escopo_final['Novo SLO Local'] = df_escopo_final['Novo SLO Local'].apply(lambda x: x if pd.notnull(x) and x > 0 else 0).astype(int)
                    colunas_finais_abrangencia = ['Cidade', 'State', 'Região de preço', 'Novo SLO Local', 'Observação']
                    
                    regioes_finais_destino = sorted(df_escopo_final['Região de preço'].unique().tolist())
                    
                    # CÁLCULO DE CONTEXTO (Seção 8)
                    context_data = []
                    for reg in regioes_finais_destino:
                        cidades_reg = df_escopo_final[df_escopo_final['Região de preço'] == reg]
                        bases_envolvidas = df_volume[df_volume['Cidade'].isin(cidades_reg['Cidade'])]['Leve'].unique()
                        
                        for base in bases_envolvidas:
                            vols = df_volume[(df_volume['Leve'] == base) & (df_volume['Cidade'].isin(cidades_reg['Cidade']))]
                            vol_total = vols['# Total Packages'].sum()
                            custo_total = 0
                            
                            tb = df_frete_clean[(df_frete_clean['LMC name'] == base) & (df_frete_clean['label'] == reg)]
                            fx1 = tb[tb['Faixa de peso cubado (g)'].str.contains('01 0 to 300', na=False, case=False)]
                            val_fx1 = fx1['on time amount'].values[0] if not fx1.empty else 0
                            
                            for _, v_row in vols.iterrows():
                                qtd = v_row['# Total Packages']
                                fx = v_row['Faixa de peso cubado (g)']
                                tb_fx = tb[tb['Faixa de peso cubado (g)'] == fx]
                                preco = tb_fx['on time amount'].values[0] if not tb_fx.empty else 0
                                custo_total += qtd * preco
                                
                            tk_medio = custo_total / vol_total if vol_total > 0 else 0
                            if vol_total > 0:
                                context_data.append({
                                    "Região de Preço": reg,
                                    "Base (Origem/Atual)": base,
                                    "Volumetria (30d)": int(vol_total),
                                    "Valor 1ª Faixa Atual": formatar_moeda(val_fx1),
                                    "Ticket Médio Atual": formatar_moeda(tk_medio)
                                })
                    
                    with st.expander("8. Ajustes Comerciais e Cenários", expanded=True):
                        st.markdown("### ℹ️ Contexto Atual das Regiões Envolvidas")
                        if context_data:
                            st.dataframe(pd.DataFrame(context_data), hide_index=True, use_container_width=True)
                        else:
                            st.info("Nenhuma volumetria recente encontrada para as cidades selecionadas.")
                            
                        c_head, c_btn = st.columns([5, 1.5])
                        with c_head:
                            st.markdown("### 🎛️ Configuração de Cenários")
                        with c_btn:
                            st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
                            if st.button("➕ Novo Cenário", use_container_width=True):
                                st.session_state["num_cenarios"] += 1
                                st.rerun()

                        cenarios_nomes = [f"Cenário {i+1}" for i in range(st.session_state["num_cenarios"])]
                        
                        css_tabs = "<style>"
                        for i in range(st.session_state["num_cenarios"]):
                            bg_cor = CORES_CENARIOS[i % len(CORES_CENARIOS)]
                            css_tabs += f'button[data-baseweb="tab"]:nth-child({i+1}) {{ background-color: {bg_cor}; border-radius: 6px 6px 0 0; margin-right: 2px; border: 1px solid #ccc; border-bottom: none; }}\n'
                        css_tabs += "</style>"
                        st.markdown(css_tabs, unsafe_allow_html=True)

                        tabs_cenarios = st.tabs(cenarios_nomes)
                        
                        for c_idx, tab in enumerate(tabs_cenarios):
                            cen_id = f"c{c_idx+1}"
                            with tab:
                                st.markdown(f"**Ajuste em Massa - {cenarios_nomes[c_idx]}**")
                                cg1, cg2, cg3, cg4 = st.columns([1.5, 1.5, 2, 3])
                                with cg1: st.selectbox("Tipo de Ajuste", ["%", "R$ (1ª Faixa)", "R$ (Ticket Médio)"], key=f"global_tipo_{cen_id}")
                                with cg2: st.number_input("Valor", step=0.5, format="%.2f", key=f"global_val_{cen_id}")
                                with cg3:
                                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                                    def aplicar_global(c_id=cen_id):
                                        t = st.session_state[f"global_tipo_{c_id}"]
                                        v = st.session_state[f"global_val_{c_id}"]
                                        for r in regioes_finais_destino:
                                            st.session_state[f"tipo_{c_id}_{r}"] = t
                                            st.session_state[f"val_{c_id}_{r}"] = v
                                    st.button("Aplicar a todas", key=f"btn_glob_{cen_id}", on_click=aplicar_global)
                                
                                st.divider()
                                st.markdown("**Ajustes Individuais:**")
                                cols_ajuste = st.columns(3)
                                for i, regiao in enumerate(regioes_finais_destino):
                                    if f"tipo_{cen_id}_{regiao}" not in st.session_state: st.session_state[f"tipo_{cen_id}_{regiao}"] = "%"
                                    if f"val_{cen_id}_{regiao}" not in st.session_state: st.session_state[f"val_{cen_id}_{regiao}"] = 0.0
                                    
                                    with cols_ajuste[i % 3]:
                                        st.markdown(f"**{regiao}**")
                                        c_tipo, c_val, c_btn = st.columns([2, 2, 1.5])
                                        with c_tipo: st.selectbox("Tipo", ["%", "R$ (1ª Faixa)", "R$ (Ticket Médio)"], key=f"tipo_{cen_id}_{regiao}", label_visibility="collapsed")
                                        with c_val: st.number_input("Valor", step=0.5, format="%.2f", key=f"val_{cen_id}_{regiao}", label_visibility="collapsed")
                                        with c_btn:
                                            def zerar_reg(c_id=cen_id, r=regiao):
                                                st.session_state[f"tipo_{c_id}_{r}"] = "%"
                                                st.session_state[f"val_{c_id}_{r}"] = 0.0
                                            st.button("Zerar", key=f"btn_zerar_{cen_id}_{regiao}", on_click=zerar_reg, use_container_width=True)
                                        st.markdown("<br>", unsafe_allow_html=True)

                    with st.expander("9. Resumo Comparativo de Cenários", expanded=True):
                        st.success(f"Cálculos finalizados para **{nome_destino_final}**!")
                        
                        resultados_cenarios = []
                        cenario_metrics = {}
                        
                        for c_idx, cenario_nome in enumerate(cenarios_nomes):
                            cen_id = f"c{c_idx+1}"
                            lista_tabelas_regiao = []
                            
                            fat_atual_total = 0
                            fat_simulado_total = 0
                            vol_total_cenario = 0
                            
                            for regiao in regioes_finais_destino:
                                cidades_mov_regiao = df_movidos[df_movidos['Região de preço'] == regiao]
                                custo_alvo_on = 0
                                soma_vol_mult = 0
                                vol_regiao = 0
                                custo_atual_regiao = 0
                                
                                mult_dict = dict(zip(df_price_var_clean['Faixa de peso cubado (g)'], df_price_var_clean['Multiplicador']))
                                
                                def calc_base_vol(cidades_df, lmc_col):
                                    v = 0; c_on = 0; s_mult = 0; c_atual = 0
                                    for _, row in cidades_df.iterrows():
                                        leve_orig = row[lmc_col]
                                        cid = row['Cidade']
                                        vols = df_volume[(df_volume['Leve'] == leve_orig) & (df_volume['Cidade_Normalizada'] == normalize_string(cid))]
                                        for _, v_row in vols.iterrows():
                                            fx = v_row['Faixa de peso cubado (g)']
                                            qtd = v_row['# Total Packages']
                                            tb_antiga = df_frete_clean[(df_frete_clean['LMC name'] == leve_orig) & (df_frete_clean['label'] == regiao) & (df_frete_clean['Faixa de peso cubado (g)'] == fx)]
                                            p_on = tb_antiga['on time amount'].values[0] if not tb_antiga.empty else 0
                                            
                                            v += qtd
                                            c_atual += qtd * p_on
                                            c_on += qtd * p_on
                                            s_mult += qtd * mult_dict.get(fx, 1.0)
                                    return v, c_on, s_mult, c_atual

                                if regiao in df_movidos['Região de preço'].unique() and estrategia_preco == "Gerar Tabela Equivalente (Focada em manter Impacto Neutro)":
                                    v1, co1, sm1, ca1 = calc_base_vol(cidades_mov_regiao, 'LMC Name')
                                    vol_regiao += v1; custo_alvo_on += co1; soma_vol_mult += sm1; custo_atual_regiao += ca1
                                    
                                    if tipo_destino == "Um Leve Existente (já selecionado)":
                                        cidades_exist = df_abrangencia_existente[df_abrangencia_existente['Região de preço'] == regiao]
                                        cidades_exist['LMC Dummy'] = nome_destino_final
                                        v2, co2, sm2, ca2 = calc_base_vol(cidades_exist, 'LMC Dummy')
                                        vol_regiao += v2; custo_alvo_on += co2; soma_vol_mult += sm2; custo_atual_regiao += ca2
                                        
                                    base_on = custo_alvo_on / soma_vol_mult if soma_vol_mult > 0 else 0
                                else:
                                    frete_base = df_frete_clean[(df_frete_clean['LMC name'] == tabela_base_selecionada) & (df_frete_clean['label'] == regiao)]
                                    faixa1_base = frete_base[frete_base['Faixa de peso cubado (g)'].str.contains('01 0 to 300', na=False, case=False)]
                                    nova_faixa1_on = faixa1_base['on time amount'].values[0] if not faixa1_base.empty else 0
                                    base_on = nova_faixa1_on / 0.83
                                    
                                    v1, co1, sm1, ca1 = calc_base_vol(cidades_mov_regiao, 'LMC Name')
                                    vol_regiao += v1; custo_atual_regiao += ca1; soma_vol_mult += sm1
                                    if tipo_destino == "Um Leve Existente (já selecionado)":
                                        cidades_exist = df_abrangencia_existente[df_abrangencia_existente['Região de preço'] == regiao]
                                        cidades_exist['LMC Dummy'] = nome_destino_final
                                        v2, co2, sm2, ca2 = calc_base_vol(cidades_exist, 'LMC Dummy')
                                        vol_regiao += v2; custo_atual_regiao += ca2; soma_vol_mult += sm2

                                df_regiao_tabela = df_price_var_clean.copy()
                                df_regiao_tabela['Região de Preço'] = regiao
                                df_regiao_tabela['Valor dentro do prazo'] = df_regiao_tabela['Multiplicador'] * base_on
                                df_regiao_tabela['Valor fora do prazo'] = df_regiao_tabela['Multiplicador'] * base_on
                                
                                ajuste_tipo = st.session_state.get(f"tipo_{cen_id}_{regiao}", "%")
                                ajuste_val = st.session_state.get(f"val_{cen_id}_{regiao}", 0.0)
                                ajuste_perc = 0.0
                                
                                if ajuste_val != 0.0:
                                    if ajuste_tipo == "R$ (1ª Faixa)":
                                        val_fx1_atual = base_on * 0.83
                                        if val_fx1_atual > 0: ajuste_perc = ((ajuste_val / val_fx1_atual) - 1) * 100
                                    elif ajuste_tipo == "R$ (Ticket Médio)":
                                        tk_projetado_base = (base_on * soma_vol_mult) / vol_regiao if vol_regiao > 0 else 0
                                        if tk_projetado_base > 0: ajuste_perc = ((ajuste_val / tk_projetado_base) - 1) * 100
                                    else:
                                        ajuste_perc = ajuste_val
                                        
                                    fator = 1 + (ajuste_perc / 100)
                                    df_regiao_tabela['Valor dentro do prazo'] *= fator
                                    df_regiao_tabela['Valor fora do prazo'] *= fator
                                    
                                lista_tabelas_regiao.append(df_regiao_tabela[['Região de Preço', 'Faixa de peso cubado (g)', 'Valor dentro do prazo', 'Valor fora do prazo']])
                                
                                custo_simulado_regiao = 0
                                for _, prow in df_regiao_tabela.iterrows():
                                    fx = prow['Faixa de peso cubado (g)']
                                    p_novo = prow['Valor dentro do prazo']
                                    
                                    def count_vol(cidades_df, lmc_col):
                                        vol_fx = 0
                                        for _, r in cidades_df.iterrows():
                                            vols = df_volume[(df_volume['Leve'] == r[lmc_col]) & (df_volume['Cidade_Normalizada'] == normalize_string(r['Cidade'])) & (df_volume['Faixa de peso cubado (g)'] == fx)]
                                            vol_fx += vols['# Total Packages'].sum() if not vols.empty else 0
                                        return vol_fx
                                        
                                    v_fx = count_vol(cidades_mov_regiao, 'LMC Name')
                                    if tipo_destino == "Um Leve Existente (já selecionado)":
                                        cidades_exist = df_abrangencia_existente[df_abrangencia_existente['Região de preço'] == regiao]
                                        cidades_exist['LMC Dummy'] = nome_destino_final
                                        v_fx += count_vol(cidades_exist, 'LMC Dummy')
                                        
                                    custo_simulado_regiao += v_fx * p_novo
                                
                                fat_atual_total += custo_atual_regiao
                                fat_simulado_total += custo_simulado_regiao
                                vol_total_cenario += vol_regiao
                                
                                resultados_cenarios.append({
                                    "Cenário": cenario_nome,
                                    "Região de Preço": regiao,
                                    "Volumetria": vol_regiao,
                                    "Faturamento Atual": custo_atual_regiao,
                                    "Ticket Médio Atual": custo_atual_regiao / vol_regiao if vol_regiao > 0 else 0,
                                    "Faturamento Projetado": custo_simulado_regiao,
                                    "Ticket Médio Projetado": custo_simulado_regiao / vol_regiao if vol_regiao > 0 else 0,
                                    "Impacto Financeiro (R$)": custo_simulado_regiao - custo_atual_regiao,
                                    "% Aumento": (custo_simulado_regiao / custo_atual_regiao - 1) if custo_atual_regiao > 0 else 0
                                })
                            
                            df_tabela_final = pd.concat(lista_tabelas_regiao, ignore_index=True)
                            df_tabela_final.rename(columns={'Faixa de peso (g/m³)': 'Faixa de peso cubado (g)'}, inplace=True)
                            dict_tabelas_finais[cenario_nome] = df_tabela_final
                            
                            resultados_cenarios.append({
                                "Cenário": cenario_nome,
                                "Região de Preço": "Total Geral",
                                "Volumetria": vol_total_cenario,
                                "Faturamento Atual": fat_atual_total,
                                "Ticket Médio Atual": fat_atual_total / vol_total_cenario if vol_total_cenario > 0 else 0,
                                "Faturamento Projetado": fat_simulado_total,
                                "Ticket Médio Projetado": fat_simulado_total / vol_total_cenario if vol_total_cenario > 0 else 0,
                                "Impacto Financeiro (R$)": fat_simulado_total - fat_atual_total,
                                "% Aumento": (fat_simulado_total / fat_atual_total - 1) if fat_atual_total > 0 else 0
                            })
                            
                            g_vol_dest = 0; g_custo_dest = 0; g_vol_migr = 0; g_custo_migr = 0; g_custo_novo = 0
                            detalhes_regioes_indiv = {}
                            for regiao in regioes_finais_destino:
                                cidades_mov_regiao = df_movidos[df_movidos['Região de preço'] == regiao]
                                def calc_costs(cidades_df, lmc_col):
                                    v = 0; c_ant = 0; c_nov = 0
                                    tb_nova = dict_tabelas_finais[cenario_nome]
                                    for _, r in cidades_df.iterrows():
                                        vols = df_volume[(df_volume['Leve'] == r[lmc_col]) & (df_volume['Cidade_Normalizada'] == normalize_string(r['Cidade']))]
                                        for _, v_row in vols.iterrows():
                                            fx = v_row['Faixa de peso cubado (g)']
                                            qtd = v_row['# Total Packages']
                                            tb_antiga = df_frete_clean[(df_frete_clean['LMC name'] == r[lmc_col]) & (df_frete_clean['label'] == regiao) & (df_frete_clean['Faixa de peso cubado (g)'] == fx)]
                                            preco_ant = tb_antiga['on time amount'].values[0] if not tb_antiga.empty else 0
                                            tb_n = tb_nova[tb_nova['Faixa de peso cubado (g)'] == fx]
                                            preco_nov = tb_n['Valor dentro do prazo'].values[0] if not tb_n.empty else 0
                                            v += qtd; c_ant += qtd * preco_ant; c_nov += qtd * preco_nov
                                    return v, c_ant, c_nov
                                v_m, c_a_m, c_n_m = calc_costs(cidades_mov_regiao, 'LMC Name')
                                g_vol_migr += v_m; g_custo_migr += c_a_m; g_custo_novo += c_n_m
                                reg_vol = v_m; reg_c_antigo = c_a_m; reg_c_novo = c_n_m
                                if tipo_destino == "Um Leve Existente (já selecionado)":
                                    cidades_exist = df_abrangencia_existente[df_abrangencia_existente['Região de preço'] == regiao]
                                    cidades_exist['LMC Dummy'] = nome_destino_final
                                    v_d, c_a_d, c_n_d = calc_costs(cidades_exist, 'LMC Dummy')
                                    g_vol_dest += v_d; g_custo_dest += c_a_d; g_custo_novo += c_n_d
                                    reg_vol += v_d; reg_c_antigo += c_a_d; reg_c_novo += c_n_d
                                detalhes_regioes_indiv[regiao] = {'vol': reg_vol, 'custo_antigo': reg_c_antigo, 'custo_novo': reg_c_novo, 'ajuste': st.session_state.get(f"val_{cen_id}_{regiao}", 0.0)}

                            cenario_metrics[cenario_nome] = {
                                'fat_antigo': g_custo_dest,
                                'vol_fat_antigo': g_vol_dest,
                                'tk_fat_antigo': g_custo_dest / g_vol_dest if g_vol_dest > 0 else 0,
                                'fat_novo': g_custo_novo,
                                'vol_fat_novo': g_vol_dest + g_vol_migr,
                                'tk_fat_novo': g_custo_novo / (g_vol_dest + g_vol_migr) if (g_vol_dest + g_vol_migr) > 0 else 0,
                                'cresc_fat': g_custo_novo - g_custo_dest,
                                'perc_cresc': ((g_custo_novo - g_custo_dest) / g_custo_dest * 100) if g_custo_dest > 0 else 0,
                                'loggi_antigo': g_custo_dest + g_custo_migr,
                                'vol_loggi': g_vol_dest + g_vol_migr,
                                'tk_loggi_antigo': (g_custo_dest + g_custo_migr) / (g_vol_dest + g_vol_migr) if (g_vol_dest + g_vol_migr) > 0 else 0,
                                'loggi_novo': g_custo_novo,
                                'tk_loggi_novo': g_custo_novo / (g_vol_dest + g_vol_migr) if (g_vol_dest + g_vol_migr) > 0 else 0,
                                'imp_loggi': g_custo_novo - (g_custo_dest + g_custo_migr),
                                'perc_imp_loggi': ((g_custo_novo - (g_custo_dest + g_custo_migr)) / (g_custo_dest + g_custo_migr) * 100) if (g_custo_dest + g_custo_migr) > 0 else 0,
                                'detalhes_regioes': detalhes_regioes_indiv
                            }

                        df_res_bruto = pd.DataFrame(resultados_cenarios)
                        df_base_atual = df_res_bruto[df_res_bruto['Cenário'] == cenarios_nomes[0]][['Região de Preço', 'Volumetria', 'Faturamento Atual', 'Ticket Médio Atual']]
                        df_comparativo = df_base_atual.copy()
                        
                        for cen in cenarios_nomes:
                            df_c = df_res_bruto[df_res_bruto['Cenário'] == cen][['Região de Preço', 'Faturamento Projetado', 'Ticket Médio Projetado', 'Impacto Financeiro (R$)', '% Aumento']]
                            df_c.columns = ['Região de Preço', f'Fat. {cen}', f'TK {cen}', f'Impacto {cen}', f'% Aum. {cen}']
                            df_comparativo = df_comparativo.merge(df_c, on='Região de Preço')
                        
                        row_total = df_comparativo[df_comparativo['Região de Preço'] == 'Total Geral']
                        df_comparativo = df_comparativo[df_comparativo['Região de Preço'] != 'Total Geral']
                        df_comparativo = pd.concat([df_comparativo, row_total], ignore_index=True)
                        
                        # --- EXIBIÇÃO EM ABAS (SEÇÃO 9) ---
                        css_tabs_resumo = "<style>"
                        for i in range(st.session_state["num_cenarios"]):
                            bg_cor = CORES_CENARIOS[i % len(CORES_CENARIOS)]
                            css_tabs_resumo += f'button[data-baseweb="tab"]:nth-child({i+2}) {{ background-color: {bg_cor}; border-radius: 6px 6px 0 0; margin-right: 2px; border: 1px solid #ccc; border-bottom: none; }}\n'
                        css_tabs_resumo += "</style>"
                        st.markdown(css_tabs_resumo, unsafe_allow_html=True)

                        tabs_res = st.tabs(["📊 Resumo Geral"] + cenarios_nomes)
                        
                        with tabs_res[0]:
                            st.markdown("### 🏆 Destaques do Cenário Atual")
                            tk_global_atual = cenario_metrics[cenarios_nomes[0]]['tk_loggi_antigo']
                            vol_global_atual = cenario_metrics[cenarios_nomes[0]]['vol_loggi']
                            
                            st.metric("Ticket Médio Atual Global", formatar_moeda(tk_global_atual), f"Volume Total: {int(vol_global_atual):,} pacotes", delta_color="off")
                            
                            st.markdown("### 🚀 Impacto dos Cenários Simulados")
                            cols_destaques = st.columns(len(cenarios_nomes))
                            for idx, cen in enumerate(cenarios_nomes):
                                m = cenario_metrics[cen]
                                with cols_destaques[idx]:
                                    st.markdown(f"**{cen}**")
                                    t_novo = formatar_moeda(m['tk_loggi_novo'])
                                    imp = m['imp_loggi']
                                    imp_perc = m['perc_imp_loggi']
                                    
                                    st.markdown(f"**Ticket Médio:** {t_novo}")
                                    if imp > 0:
                                        st.markdown(f"**Diferença Mensal:** <span style='color:#ff4b4b'>▲ +{formatar_moeda(imp)}</span>", unsafe_allow_html=True)
                                        st.markdown(f"**% Aumento:** <span style='color:#ff4b4b'>▲ +{imp_perc:.2f}%</span>", unsafe_allow_html=True)
                                    elif imp < 0:
                                        st.markdown(f"**Diferença Mensal:** <span style='color:#09ab3b'>▼ -{formatar_moeda(abs(imp))}</span>", unsafe_allow_html=True)
                                        st.markdown(f"**% Aumento:** <span style='color:#09ab3b'>▼ {imp_perc:.2f}%</span>", unsafe_allow_html=True)
                                    else:
                                        st.markdown(f"**Diferença Mensal:** <span style='color:gray'>■ R$ 0,00</span>", unsafe_allow_html=True)
                                        st.markdown(f"**% Aumento:** <span style='color:gray'>■ 0.00%</span>", unsafe_allow_html=True)
                            
                            st.divider()
                            st.markdown("### 📋 Quadro Resumo Comparativo")
                            
                            df_disp = df_comparativo.copy()
                            for c in df_disp.columns:
                                if "Fat" in c or "Ticket" in c or "TK" in c or "Impacto" in c:
                                    df_disp[c] = df_disp[c].apply(lambda x: formatar_moeda(x) if pd.notna(x) else "-")
                                elif "%" in c or "Aum" in c:
                                    df_disp[c] = df_disp[c].apply(lambda x: f"{x*100:+.2f}%" if pd.notna(x) else "-")
                                elif "Vol" in c:
                                    df_disp[c] = df_disp[c].apply(lambda x: f"{int(x):,}".replace(",", ".") if pd.notna(x) else "-")
                            
                            def get_styles(df):
                                styles = pd.DataFrame('', index=df.index, columns=df.columns)
                                for col in df.columns:
                                    for i, cen in enumerate(cenarios_nomes):
                                        if cen in col or f"Cenário {i+1}" in col:
                                            styles[col] = f'background-color: {CORES_CENARIOS[i % len(CORES_CENARIOS)]}; color: #333333; text-align: center;'
                                        else:
                                            styles[col] = 'text-align: center;'
                                
                                for idx in df.index:
                                    if df.loc[idx, 'Região de Preço'] == 'Total Geral':
                                        styles.loc[idx, :] = 'background-color: #002766; color: white; font-weight: bold; text-align: center;'
                                return styles
                                
                            styled_df = df_disp.style.apply(get_styles, axis=None)
                            st.dataframe(styled_df, hide_index=True, use_container_width=True)

                        for idx, cen in enumerate(cenarios_nomes):
                            with tabs_res[idx+1]:
                                m = cenario_metrics[cen]
                                st.subheader(f"🤝 Visão do Parceiro ({cen})")
                                cp1, cp2, cp3 = st.columns(3)
                                with cp1:
                                    st.metric("Faturamento Atual (Sem Novas Cidades)", formatar_moeda(m['fat_antigo']))
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Volumetria: {int(m['vol_fat_antigo']):,} pacotes</span>", unsafe_allow_html=True)
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Ticket Médio: {formatar_moeda(m['tk_fat_antigo'])}</span>", unsafe_allow_html=True)
                                with cp2:
                                    st.metric("Novo Faturamento Projetado", formatar_moeda(m['fat_novo']))
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Volumetria: {int(m['vol_fat_novo']):,} pacotes</span>", unsafe_allow_html=True)
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Ticket Médio: {formatar_moeda(m['tk_fat_novo'])}</span>", unsafe_allow_html=True)
                                with cp3:
                                    st.metric("Crescimento da Operação", formatar_moeda(m['cresc_fat']))
                                    st.markdown(f"<span style='font-size: 0.9em; color: #09ab3b; font-weight: bold;'>▲ +{m['perc_cresc']:.2f}% de aumento no faturamento</span>", unsafe_allow_html=True)

                                st.divider()
                                st.subheader(f"📉 Visão Loggi ({cen})")
                                cl1, cl2, cl3 = st.columns(3)
                                with cl1:
                                    st.metric("Custo Antigo Global", formatar_moeda(m['loggi_antigo']))
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Volumetria Total: {int(m['vol_loggi']):,} pacotes</span>", unsafe_allow_html=True)
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Ticket Médio Antigo: {formatar_moeda(m['tk_loggi_antigo'])}</span>", unsafe_allow_html=True)
                                with cl2:
                                    st.metric("Novo Custo Global Projetado", formatar_moeda(m['loggi_novo']))
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Volumetria Total: {int(m['vol_loggi']):,} pacotes</span>", unsafe_allow_html=True)
                                    st.markdown(f"<span style='font-size: 0.9em; color: gray;'>Ticket Médio Novo: {formatar_moeda(m['tk_loggi_novo'])}</span>", unsafe_allow_html=True)
                                with cl3:
                                    if m['imp_loggi'] > 0:
                                        st.metric("Impacto Financeiro Loggi", formatar_moeda(m['imp_loggi']), f"+{formatar_moeda(m['imp_loggi'])} (Aumento)", delta_color="inverse")
                                        st.markdown(f"<span style='font-size: 0.9em; color: #ff4b4b; font-weight: bold;'>▲ +{m['perc_imp_loggi']:.2f}% de impacto no budget</span>", unsafe_allow_html=True)
                                    elif m['imp_loggi'] < 0:
                                        st.metric("Impacto Financeiro Loggi", formatar_moeda(m['imp_loggi']), f"{formatar_moeda(m['imp_loggi'])} (Economia)", delta_color="normal")
                                        st.markdown(f"<span style='font-size: 0.9em; color: #09ab3b; font-weight: bold;'>▼ {m['perc_imp_loggi']:.2f}% de economia no budget</span>", unsafe_allow_html=True)
                                    else:
                                        st.metric("Impacto Financeiro Loggi", "R$ 0,00", "Neutro")
                                        st.markdown(f"<span style='font-size: 0.9em; color: gray;'>0.00%</span>", unsafe_allow_html=True)
                                
                                if m['detalhes_regioes']:
                                    st.markdown(f"#### 🔍 Detalhamento das Regiões ({cen})")
                                    for reg, dados in m['detalhes_regioes'].items():
                                        st.markdown(f"##### 📍 {reg}")
                                        ajuste_aplicado = dados.get('ajuste', 0.0)
                                        if ajuste_aplicado != 0.0:
                                            st.markdown(f"<span style='font-size: 0.9em; color: #e67e22; font-weight: bold;'>Aviso: Ajuste Comercial Aplicado. Verificar Tabela Projetada.</span>", unsafe_allow_html=True)
                                        cd1, cd2, cd3 = st.columns(3)
                                        tk_r_ant = dados['custo_antigo'] / dados['vol'] if dados['vol'] > 0 else 0
                                        tk_r_nov = dados['custo_novo'] / dados['vol'] if dados['vol'] > 0 else 0
                                        imp_r = dados['custo_novo'] - dados['custo_antigo']
                                        perc_r = (imp_r / dados['custo_antigo']) * 100 if dados['custo_antigo'] > 0 else 0
                                        with cd1:
                                            st.markdown(f"**Custo Antigo:** {formatar_moeda(dados['custo_antigo'])}")
                                            st.markdown(f"<span style='font-size: 0.8em; color: gray;'>Vol: {int(dados['vol']):,} | Tk: {formatar_moeda(tk_r_ant)}</span>", unsafe_allow_html=True)
                                        with cd2:
                                            st.markdown(f"**Novo Custo:** {formatar_moeda(dados['custo_novo'])}")
                                            st.markdown(f"<span style='font-size: 0.8em; color: gray;'>Vol: {int(dados['vol']):,} | Tk: {formatar_moeda(tk_r_nov)}</span>", unsafe_allow_html=True)
                                        with cd3:
                                            color = "gray"
                                            if imp_r > 0: color = "#ff4b4b"
                                            elif imp_r < 0: color = "#09ab3b"
                                            st.markdown(f"**Variação:** <span style='color: {color}; font-weight: bold;'>{formatar_moeda(imp_r)} ({perc_r:+.2f}%)</span>", unsafe_allow_html=True)

                        st.divider()
                        st.subheader("Tabelas de Frete Peso e Abrangência")
                        st.dataframe(df_escopo_final[colunas_finais_abrangencia], hide_index=True, use_container_width=True)
                        for cen, df_tab in dict_tabelas_finais.items():
                            st.markdown(f"**Tabela Projetada - {cen}**")
                            df_exibicao_tabela = df_tab.copy()
                            df_exibicao_tabela['Valor dentro do prazo'] = df_exibicao_tabela['Valor dentro do prazo'].apply(formatar_moeda)
                            df_exibicao_tabela['Valor fora do prazo'] = df_exibicao_tabela['Valor fora do prazo'].apply(formatar_moeda)
                            st.dataframe(df_exibicao_tabela, hide_index=True, use_container_width=True)

                        st.divider()
                        
                        col_dw1, col_down3 = st.columns([1, 1])
                        with col_dw1:
                            st.markdown("### 📥 Download da Proposta em Excel")
                            st.markdown("Planilha completa com as abas de todos os cenários projetados.")
                            output = io.BytesIO()
                            colunas_finais_abrangencia_excel = [c for c in colunas_finais_abrangencia if c != 'State']
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                df_comparativo.to_excel(writer, sheet_name='Resumo de Cenários', index=False)
                                df_escopo_final[colunas_finais_abrangencia_excel].to_excel(writer, sheet_name='Abrangência e Prazos', index=False)
                                for cenario_nome, df_tab in dict_tabelas_finais.items():
                                    df_tab.to_excel(writer, sheet_name=cenario_nome[:30], index=False)
                                formatar_excel(writer)
                            
                            st.download_button(
                                label="Baixar Proposta (Excel)",
                                data=output.getvalue(),
                                file_name=f"Proposta_Cenarios_{nome_destino_final.replace(' ', '_')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary"
                            )
                                
                        with col_down3:
                            st.markdown("### 📄 Relatório Executivo")
                            st.markdown("Resumo em PDF com o quadro comparativo e tabelas.")
                            
                            if HAS_PDF_GENERATOR:
                                if not df_movidos.empty:
                                    cidades_movimentadas_str = ", ".join(sorted(df_movidos['Cidade'].str.title().unique().tolist()))
                                else:
                                    cidades_movimentadas_str = "Nenhum (Apenas reajuste da carteira atual)"

                                pdf_data = generate_html_pdf(
                                    nome_destino_final, estrategia_preco, cidades_movimentadas_str, 
                                    df_comparativo, cenario_metrics, 
                                    df_escopo_final[colunas_finais_abrangencia], 
                                    dict_tabelas_finais, tabelas_atuais_pdf, cenarios_nomes
                                )
                                st.download_button(
                                    label="Baixar Relatório (PDF)",
                                    data=pdf_data,
                                    file_name=f"Relatorio_Cenarios_{nome_destino_final.replace(' ', '_')}.pdf",
                                    mime="application/pdf",
                                    type="primary"
                                )
                            else:
                                st.warning("Instale a biblioteca 'weasyprint' (pip install weasyprint) para liberar a exportação em PDF.")

else:
    st.info("Por favor, faça o upload de **todas as 4 bases** na barra lateral para prosseguir.")